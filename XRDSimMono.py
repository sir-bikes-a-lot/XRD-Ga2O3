# -*- coding: utf-8 -*-

import openpyxl as xl
# import math
# import cmath
import numpy as np
import matplotlib.pyplot as plt
import matplotlib as mpl

# Get the parameters for a specified binary compound and return as a
# dictionary.
def binaries(binary):
    # Load the binary compounds workbook and get the corresponding sheet.
    wb = xl.load_workbook('binaries.xlsx')
    try:
        sheet = wb[binary]
        # Get the list of the binary compound parameters.
        values = []
        for row in range(1, sheet.max_row + 1):
            cell = sheet.cell(row, 2)
            values.append(cell.value)
        # Stuff the parameters into a dictionary and return them.
        params = {
            'a': values[0],
            'b': values[1],
            'c': values[2],
            'beta': values[3],
            'C11': values[4],
            'C12': values[5],
            'C13': values[6],
            'C15': values[7],
            'C22': values[8],
            'C23': values[9],
            'C25': values[10],
            'C33': values[11],
            'C35': values[12],
            'C55': values[13]}
        return params
    except KeyError:
        print('Invalid binary compound specified.')


def elements(element):
    wb = xl.load_workbook('elements.xlsx')
    try:
        sheet = wb[element]
        # Get the list of the elemental parameters.
        values = []
        for row in range(1, sheet.max_row + 1):
            cell = sheet.cell(row, 2)
            values.append(cell.value)
        # Stuff the Cromer-Mann coefficients into a dictionary.
        CromerMann = {
            'a1': values[0],
            'a2': values[1],
            'a3': values[2],
            'a4': values[3],
            'b1': values[4],
            'b2': values[5],
            'b3': values[6],
            'b4': values[7],
            'c': values[8]}
        # Nest the Cromer-Mann coefficients into the element dictionary and
        # return it.
        params = {
            'CromerMann': CromerMann,
            'fprime': values[9],
            'f2prime': values[10],
            'A': values[11],
            'mu': values[12],
            'B': values[13]}
        return params
    except KeyError:
        print('Invalid element specified.')


# Pull in the binary compound libraries as a global variable.
wb = xl.load_workbook('binaries.xlsx')
names = wb.sheetnames
Binaries = {}
for name in names:
    Binaries[name] = binaries(name)

# Pull in the element libraries as a global variable.
wb = xl.load_workbook('elements.xlsx')
names = wb.sheetnames
Elements = {}
for name in names:
    Elements[name] = elements(name)
del [wb, names]


# Applies Vegard's Law interpolation to the parameter specified by 'param'.
def Vegard(siteAatoms, siteBatoms, binaries, param):
    siteAnames = list(siteAatoms.keys())
    siteBnames = list(siteBatoms.keys())
    binary1 = siteAnames[2] + siteBnames[0]
    binary2 = siteAnames[1] + siteBnames[0]
    binary3 = siteAnames[0] + siteBnames[0]
    b1 = binaries[binary1][param]
    b2 = binaries[binary2][param]
    b3 = binaries[binary3][param]
    x = siteAatoms[siteAnames[0]]
    y = siteAatoms[siteAnames[1]]
    value = (1-x-y)*b1 + y*b2 + x*b3
    return value


# Calculate tilt between substrate normal (u,v,w) and lattice plane (h,k,l).
# Tilt is zero for a symmetric scan. Otherwise, it is an asymmetric scan.
def CalcTilt(orient=(0, 1, 0), plane=(0, 2, 0), offcut=0):
    # Unpack orientation and plane.
    u, v, w = orient
    h, k, l = plane
    # Calculate tilt angle phi in radians. Offcut angle also assumed in radians
    phi = np.arccos((u*h + v*k + w*l)/np.sqrt((u**2 + v**2 + w**2)*(h**2 + k**2 + l**2)))
    phi += offcut
    return phi


# Calculate the out-of-plane lattice constant for the monoclinic beta-Ga2O3
# unit cell, assuming coherent strain to the substrate.
def GetStrainedLattice(relaxed, elastic, substrate, orient, r):
    # Unpack lattice constants and elastic constants.
    a, b, c, beta = relaxed
    a0, b0, c0, beta0 = substrate
    C11, C12, C13, C15, C22, C23, C25, C33, C35, C55 = elastic
    # Consider three substrate orientations: (010), (001), and (100).
    if str(orient) == '(0, 1, 0)':
        # For coherently strained layers, the in-plane lattice parameters
        # ac, cc, and betac match the substrate.
        ac = a0
        cc = c0
        betac = beta0
        # Calculate the four strains.
        eps1 = -(a - ac)/a
        eps3 = cc*np.sin(betac*np.pi/180)/(c*np.sin(beta*np.pi/180)) - 1
        eps5 = cc*np.cos(betac*np.pi/180)/(c*np.sin(beta*np.pi/180)) - ac*np.cos(beta*np.pi/180)/(a*np.sin(beta*np.pi/180))
        eps2 = -(C12*eps1 + C23*eps3 + C25*eps5)/C22
        # Calculate the out-of-plane lattice parameter bc.
        bc = b*(1 + eps2)
    elif str(orient) == '(0, 0, 1)':
        # For coherently strained layers, the in-plane lattice parameters
        # ac and bc match the substrate.
        ac = a0
        bc = b0
        # Assume that betac = beta0, justified by the small mismatch
        # between Ga2)3, Al2O3, and In2O3.
        betac = beta0
        # Calculate the two of the four strains.
        eps1 = -(a - ac)/a
        eps2 = -(b - bc)/b
        # Calculate the out-of-plane lattice parameter cc.
        cc = c*(C33*np.sin(beta*np.pi/180) - (C13*eps1 + C23*eps2)*np.sin(beta*np.pi/180) + (ac/a)*C35*np.cos(beta*np.pi/180))
        cc = cc/(C33*np.sin(betac*np.pi/180) + C35*np.cos(betac*np.pi/180))
        # Calculate the other two strains. For information only
        eps5 = cc*np.cos(betac*np.pi/180)/(c*np.sin(beta*np.pi/180)) - ac*np.cos(beta*np.pi/180)/(a*np.sin(beta*np.pi/180))
        eps3 = -(C13*eps1 + C23*eps2 + C35*eps5)/C33
    elif str(orient) == '(1, 0, 0)':
        # TO-DO: Need to calculate the transformed elastic constants for
        # this orientation! Just assume the values for Ga2O3.
        C11 = 258
        C12 = 118
        C13 = 139
        C15 = -23
        # For coherently strained layers, the in-plane lattice parameters
        # bc and cc match the substrate.
        bc = b0
        cc = c0
        # Assume that betac = beta0, justified by the small mismatch
        # between Ga2)3, Al2O3, and In2O3.
        betac = beta0
        # Calculate two of the four strains.
        eps2 = -(b - bc)/b
        eps3 = -(c - cc)/c
        # Calculate the out-of-plane lattice parameter ac.
        ac = a*((C13*eps3 + C12*eps2)*np.sin(beta*np.pi/180) + (cc/c)*C15*np.cos(betac*np.pi/180) - C11*np.sin(beta*np.pi/180))
        ac = ac/(C15*np.cos(beta*np.pi/180) - C11*np.sin(betac*np.pi/180))
        # Calculate the other two strains. For information only.
        eps5 = cc*np.cos(betac*np.pi/180)/(c*np.sin(beta*np.pi/180)) - ac*np.cos(beta*np.pi/180)/(a*np.sin(beta*np.pi/180))
        eps1 = -(C13*eps3 + C12*eps2 + C15*eps5)/C11
    else:
        print('Invalid substrate orientation. Must be (0, 1, 0), (0, 0, 1), or (1, 0, 0).')
    # Apply relaxation.
    ac = (1-r)*ac + r*a
    bc = (1-r)*bc + r*b
    cc = (1-r)*cc + r*c
    betac = (1-r)*betac + r*beta
    # Pack up the coherently strained lattice parameters.
    strained = (ac, bc, cc, betac)
    strains = (eps1, eps2, eps3, eps5)
    # Return only the strained lattice constants.
    return strained


# Calculate the Bragg angle for the specified diffraction plane.
def GetBraggAngle(a, b, c, beta, lam, plane):
    h, k, l = plane
    # Calculate interplanar spacing using auxiliary variables.
    term1 = (h/a)**2
    term2 = (k/b)**2*(np.sin(beta*np.pi/180))**2
    term3 = (l/c)**2
    term4 = -2*h*l*np.cos(beta*np.pi/180)/(a*c)
    invdsq = (term1 + term2 + term3 + term4)/((np.sin(beta*np.pi/180))**2)
    invd = np.sqrt(invdsq)
    # Calculate the m-th Bragg angle.
    thetaB = np.arcsin(lam*invd/2)
    return thetaB


# Calculates the atomic scattering factor f0 using the Cromer-Mann expression.
def CromerMann(coeffs, thetaB, lam):
    f0 = 0
    k = np.sin(thetaB)/lam
    for j in range(1, 5):
        field1 = 'a' + str(j)
        field2 = 'b' + str(j)
        f0 += coeffs[field1]*np.exp(-coeffs[field2]*(k**2))
    f0 += coeffs['c']
    return f0


# CalcStructFact calculates the x-ray structure factor for a monoclinic
# beta-Ga2O3 crystal with group III atoms at site A (0.16,0.50,0.31) and
# (0.09,0,0.79) and group VI atoms at site B (0.50,0,0.26), (0.17,0,0.56),
# and (0.34,0.50,0.89).
def CalcStructFact(fA, fB, plane):
    h, k, l = plane
    F1 = fA*np.exp(-2*np.pi*1j*(h*0.16 + k*0.50 + l*0.31))
    F2 = fA*np.exp(-2*np.pi*1j*(h*0.09 + k*0.00 + l*0.79))
    F3 = fB*np.exp(-2*np.pi*1j*(h*0.50 + k*0.00 + l*0.26))
    F4 = fB*np.exp(-2*np.pi*1j*(h*0.17 + k*0.00 + l*0.56))
    F5 = fB*np.exp(-2*np.pi*1j*(h*0.34 + k*0.50 + l*0.89))
    F = F1 + F2 + F3 + F4 + F5
    return F


# Calculate x-ray structure factor for compound. Ternary and quaternary
# compounds are handled by first calculating the atomic scattering factors
#  for the constituent binaries, then using Vegard's law to interpolate
# between them.
def GetStructFact(siteAatoms, siteBatoms, elements, thetaB, lam, plane):
    # First we need to calculate the atomic scattering factors for each atom in
    # the compound. Begin by finding f0 from the Cromer-Mann coefficients.
    siteAnames = list(siteAatoms.keys())
    siteBnames = list(siteBatoms.keys())
    atomnames = siteAnames + siteBnames
    # Initialize some variables before entering the loop.
    f0 = np.zeros(len(atomnames))
    f = np.zeros(len(atomnames), dtype=complex)
    # Fetch the Cromer-Mann coefficients from the database of elements. Then,
    # use the Cromer-Mann formula to calculate f0.
    element = []
    coeffs = []
    for j in range(len(atomnames)):
        element.append(elements[atomnames[j]])
        coeffs.append(element[j]['CromerMann'])
        f0[j] = CromerMann(coeffs[j], thetaB, lam)
        # Now add the dispersion correction factors to get atomic scattering
        # factors, f.
        f[j] = f0[j] + element[j]['fprime'] + 1j*element[j]['f2prime']
        # DEBUG: Finally, apply the Debye-Waller factor to the atomic
        # scattering factors to account for mean atomic displacement due to
        # temperature & zero-point energy.
        # f[j] = f[j]*np.exp(-(np.sin(thetaB)/lam)**2*element[j]['B'])
    # Compound is a quaternary of the form A(x)B(y)C(1-x-y)D.
    # Binary1 is CD, binary2 is BD, binary3 is AD.
    # Mole fractions x and y are given by:
    x = siteAatoms[siteAnames[0]]
    y = siteAatoms[siteAnames[1]]
    # Apply Vegard's Law to get the quaternary atomic scattering factors.
    fa = x*f[0] + y*f[1] + (1-x-y)*f[2]
    # Calculate the x-ray structure factor.
    F = CalcStructFact(fa, f[3], plane)
    return F


# Calculate the total linear absorption coefficient mu for the layer. Use
# Vegard's Law to interpolate the mass absorption coefficients.
def GetMu(siteAatoms, siteBatoms, elements, n, V):
    N = 6.022 * (10**23)
    V = V/(10**24)
    # First we need to fetch the mass absorption coefficients mu/rho for each
    # element in the compound.
    siteAnames = list(siteAatoms.keys())
    siteBnames = list(siteBatoms.keys())
    atomnames = siteAnames + siteBnames
    # Initialize mu_roh and atomic mass A before entering the loop.
    mu_rho = np.zeros(len(atomnames))
    A = np.zeros(len(atomnames))
    mu_a = np.zeros(len(atomnames))
    # Fetch mu/rho and atomic mass A for each element. Calculate atomic
    # absorption coefficients mu_a.
    element = []
    for j in range(len(atomnames)):
        element.append(elements[atomnames[j]])
        mu_rho[j] = element[j]['mu']
        A[j] = element[j]['A']
        mu_a[j] = mu_rho[j]*A[j]/N
    # Compound is a quaternary of the form A(x)B(y)C(1-x-y)D.
    # Binary1 is CD, binary2 is BD, binary3 is AD.
    # Mole fractions x and y are given by:
    x = siteAatoms[siteAnames[0]]
    y = siteAatoms[siteAnames[1]]
    # Apply Vegard's Law to the group III atomic absorption coefficients.
    mu = (1-x-y)*mu_a[2] + y*mu_a[1] + x*mu_a[0]
    # Now add the group V atomic absorption coefficient.
    mu += mu_a[3]
    mu = (n/V)*mu
    return mu


# Set up an output structure for each of the N lamellae in the sample stackup
# and the specified XRD scan range (in radians).
def ProcessLamella(sample, Binaries, Elements, omega, phi=0, delpsi=0,
                   plane=(0, 2, 0), ScanType='Omega-2Theta', Twotheta=0,
                   lam=1.540594, a0=12.214, b0=3.0371, c0=5.7981, beta0=103.83,
                   orient=(0, 1, 0), thetaM=0.3953):
    # Apply Vegard's Law to find the lattice constants a, b, c, and beta, and
    # the elastic constants.
    a = Vegard(sample['siteAatoms'], sample['siteBatoms'], Binaries, 'a')
    b = Vegard(sample['siteAatoms'], sample['siteBatoms'], Binaries, 'b')
    c = Vegard(sample['siteAatoms'], sample['siteBatoms'], Binaries, 'c')
    beta = Vegard(sample['siteAatoms'], sample['siteBatoms'], Binaries, 'beta')
    C11 = Vegard(sample['siteAatoms'], sample['siteBatoms'], Binaries, 'C11')
    C12 = Vegard(sample['siteAatoms'], sample['siteBatoms'], Binaries, 'C12')
    C13 = Vegard(sample['siteAatoms'], sample['siteBatoms'], Binaries, 'C13')
    C15 = Vegard(sample['siteAatoms'], sample['siteBatoms'], Binaries, 'C15')
    C22 = Vegard(sample['siteAatoms'], sample['siteBatoms'], Binaries, 'C22')
    C23 = Vegard(sample['siteAatoms'], sample['siteBatoms'], Binaries, 'C23')
    C25 = Vegard(sample['siteAatoms'], sample['siteBatoms'], Binaries, 'C25')
    C33 = Vegard(sample['siteAatoms'], sample['siteBatoms'], Binaries, 'C33')
    C35 = Vegard(sample['siteAatoms'], sample['siteBatoms'], Binaries, 'C35')
    C55 = Vegard(sample['siteAatoms'], sample['siteBatoms'], Binaries, 'C55')
    # Pack up the lattice constants and elastic constants.
    substrate = (a0, b0, c0, beta0)
    relaxed = (a, b, c, beta)
    elastic = (C11, C12, C13, C15, C22, C23, C25, C33, C35, C55)
    # Calculate coherently strained lattice constants.
    strained = GetStrainedLattice(relaxed, elastic, substrate, orient, sample['relaxation'])
    # Unpack the coherently strained lattice parameters.
    ac, bc, cc, betac = strained
    # If this is an interface offset layer, manually adjust the lattice
    # parameter perpendicular to the substrate by the offset.
    # if (strcmp(sample.layertype, 'offset')):
    if sample['layertype'] == 'offset':
        if orient[0] == 1:
            ac = ac - sample['thickness']
            sample['thickness'] = a0
        elif orient[1] == 1:
            bc = bc - sample['thickness']
            sample['thickness'] = b0
        elif orient[2] == 1:
            cc = cc - sample['thickness']
            sample['thickness'] = c0
    # Calculate Bragg angle for the strained unit cell. Specify the m-th Bragg
    # reflection. Bragg angle is in radians.
    thetaB = GetBraggAngle(ac, bc, cc, betac, lam, plane)
    # Calculate unitless Gamma parameter.
    re = 2.818*10**(-5)                       # Classical electron radius in Angstroms.
    V = ac*bc*cc*np.sin(betac*np.pi/180)  # Unit cell volume in Angstroms.
    Gamma = (re*lam**2)/(np.pi*V)           # Unitless parameter.
    # Calculate theta, the angle between the incident X-ray beam and the
    # lattice plane (h,k,l). This is offset by the scan tilt.
    delphi = sample['delphi']
    theta = omega - phi
    # Calculate structure factors for (000), (hkl), and (-h-k-l) planes.
    F0 = GetStructFact(sample['siteAatoms'], sample['siteBatoms'], Elements, thetaB, lam, (0, 0, 0))
    FH = GetStructFact(sample['siteAatoms'], sample['siteBatoms'], Elements, thetaB, lam, plane)
    FHbar = GetStructFact(sample['siteAatoms'], sample['siteBatoms'], Elements, thetaB, lam, [-x for x in plane])
    # Calculate total linear absorption coefficient for x-ray absorption.
    mu = GetMu(sample['siteAatoms'], sample['siteBatoms'], Elements, 6, V)
    # Calculate X-ray polarization factor, C. Random polarization corresponds
    # to thetaM = 0.
    C = (1 + (np.cos(2*thetaM))**2*(np.cos(2*thetaB))**2)/(1 + (np.cos(2*thetaM))**2)
    # Calculate unitless thickness parameter.
    h = 10*sample['thickness']
    gamma0 = np.sin(omega)
    # Direction cosine of diffracted beam depends on scan type.
    if ScanType == 'Rocking':
        gammaH = -np.sin(omega - Twotheta)
    elif ScanType == 'Omega-2Theta':
        gammaH = gamma0
    else:
        print('Invalid scan type. Must be either "Rocking" or "Omega-2Theta"')
    T = np.zeros(len(theta), dtype=complex)
    for i in range(len(theta)):
        T[i] = h*C*(np.pi*Gamma*np.sqrt(FH*FHbar))/(lam*np.sqrt(abs(gamma0[i]*gammaH[i])))
    # Setup output structure "lamella".
    lamella = {
        'Theta': theta,
        'delphi': delphi,
        'delpsi': delpsi,
        'BraggAngle': thetaB,
        'sigma_alpha': sample['sigma_alpha'],
        'sigma_beta': sample['sigma_beta'],
        'Gamma': Gamma,
        'F0': F0,
        'FH': FH,
        'FHbar': FHbar,
        'mu': mu,
        'T': T,
        'gamma0': gamma0,
        'gammaH': gammaH,
        'C': C}
    return lamella


# Calculate the mosaic crystal weighting distribution based on specified
# distribution type.
def WeightDist(x, sigma, N, type):
    if type == 'Gaussian':
        if sigma <= 0:
            sigma = 1e-5
        w = []
        for y in x:
            w.append(np.exp((-y**2)/(2*sigma**2)))
#        W = [np.exp((-y**2)/(2*sigma**2)) for y in x]
#        W = [y/sum(W) for y in W]
        W = [y/sum(w) for y in w]
    elif type == 'Uniform':
        W = 1/N
    else:
        print('Invalid weighting distribution specified.')
    return W


# Calculates the deviation parameter as a function of theta, the
# fundamental quantity for determining the X-ray diffraction pattern.
def CalcEta(lamella, ang_alpha, ang_beta):
    # Unpack the lamella structure.
    theta = lamella['Theta']
    delphi = lamella['delphi']
    delpsi = lamella['delpsi']
    thetaB = lamella['BraggAngle']
    Gamma = lamella['Gamma']
    F0 = lamella['F0']
    FH = lamella['FH']
    FHbar = lamella['FHbar']
    gamma0 = lamella['gamma0']
    gammaH = lamella['gammaH']
    C = lamella['C']
    # Calculate eta in two parts, numerator & denominator.
    eta_num = -(gamma0/gammaH)*(theta - thetaB - (ang_alpha + delphi)*np.cos(ang_beta + delpsi))*np.sin(2*thetaB) - (1/2)*(1 - gamma0/gammaH)*Gamma*F0
    eta_den = C*Gamma*np.sqrt(abs(gamma0/gammaH)*FH*FHbar)
    eta = eta_num/eta_den
    return eta


# Calculate the diffracted x-ray signal amplitude for a single lamella and
# a single mosaic crystallite.
def CalcX(eta, X_last, T):
    # First, calculate the auxiliary variables S1 and S2.
    S1 = np.empty(len(eta), dtype=complex)
    S2 = np.empty(len(eta), dtype=complex)
    X = np.empty(len(eta), dtype=complex)
    for i in range(len(eta)):
        S1[i] = (X_last[i] - eta[i] + np.sqrt(eta[i]**2 - 1))*np.exp(-1j*T[i]*np.sqrt(eta[i]**2 - 1))
        S2[i] = (X_last[i] - eta[i] - np.sqrt(eta[i]**2 - 1))*np.exp(1j*T[i]*np.sqrt(eta[i]**2 - 1))
        # Now calculate diffracted signal amplitude X:
        X[i] = eta[i] + np.sqrt(eta[i]**2 - 1)*(S1[i] + S2[i])/(S1[i] - S2[i])
    return X


# Convolve simulated x-ray intensity with a normal distribution to model
# the effect of instrumental broadening.
def InstrBroaden(I, omega, FWHM):
    # Caclulate the standard deviation of the instrumental broadening normal
    # distribution.
    pts_per_rad = len(omega)/(omega[-1] - omega[0])
    width = FWHM * pts_per_rad
    sigma = width/(2*np.sqrt(2*np.log(2)))
    # Set up vector of x-values for the normal distribution. Use +/-5 standard
    # deviations.
    N = 10*sigma
    x = np.linspace(-5*sigma, 5*sigma, np.floor(N).astype(int))
    # Calculate instrumental broadening function.
    f = np.zeros(len(x))
    for i in range(len(x)):
        f[i] = (1/(np.sqrt(2*np.pi)*sigma))*np.exp((-x[i]**2)/(2*sigma**2))
    # Convolve instrumental broadening function with diffracted x-ray intensity
    # to obtain output. Trim the ends of the output to the proper length.
    I_out = np.convolve(I, f, 'same')
    return I_out


def Xray(sample, omega, orient=(0, 1, 0), plane=(0, 2, 0), offcut=0, delpsi=0,
         ScanType='Omega-2Theta', Twotheta=0, lam=1.540594, thetaM=0.3953,
         bounces=2, N_sigma=3, N_alpha=1, N_beta=1, DistAlpha='Gaussian',
         DistBeta='Uniform'):
    '''
    Simulate the diffracted intensity vs incident X-ray angle (omega).
    Input Parameters
    ----------
    sample : List of dictionaries describing the simulated sample structure.
    omega : List of incident X-ray angles, in radians.
    orient : Tuple containing the out-of-plane orientation of the substrate.
        The default is (0, 1, 0).
    plane : Tuple containing the diffraction plane for the measurement.
        The default is (0, 2, 0).
    offcut : Substrate offcut angle, in radians. The default is 0.
    delpsi : Azimuthal rotation w.r.t. offcut direction, in radians.
        The default is 0.
    ScanType : Type of scan, either 'Omega-2Theta' or 'Rocking'.
        The default is 'Omega-2Theta'.
    Twotheta : If rocking curve scan, use 2theta as detector angle in radians.
        The default is 0.
    lam : X-ray wavelength, for Cu_kalpha: 1.540594 Angstroms.
        The default is 1.540594.
    thetaM : Monochromator angle, in radians. The default is 0.3953.
    bounces : Number of monochromator bounces, usually 2 or 4. The default is 2
    N_sigma : Number of std. deviations for Gaussian distributions.
        The default is 3.
    N_alpha : Number of crystallites in tilt space. The default is 1.
    N_beta : Number of crystallites in azimuth space. The default is 1.
    DistAlpha : Type of distribution in tilt space, either "Gaussian" or
        "Uniform" The default is 'Gaussian'.
    DistBeta : Type of distribution in azimuth space, either "Gaussian" or
        "Uniform" The default is 'Uniform'.

    Returns
    -------
    I : List of simulated X-ray intensity from sample structure.
    I0 : List of simulated X-ray intensity from substrate only.
    lamella : List of dictionaries from processed sample structure.
    '''

    # Code differs here from MATLAB implementation by using Binaries and
    # Elements as global variables, rather than retrieving them from Excel
    # on every execution of XRay().

    # Get the substrate lattice constants.
    a0 = Vegard(sample[0]['siteAatoms'], sample[0]['siteBatoms'], Binaries, 'a')
    b0 = Vegard(sample[0]['siteAatoms'], sample[0]['siteBatoms'], Binaries, 'b')
    c0 = Vegard(sample[0]['siteAatoms'], sample[0]['siteBatoms'], Binaries, 'c')
    beta0 = Vegard(sample[0]['siteAatoms'], sample[0]['siteBatoms'], Binaries, 'beta')

    # Calculate tilt angle.
    phi = CalcTilt(orient, plane, offcut)

    # Loop through each layer of the sample and set up the structure for
    # ProcessLamella().
    N = len(sample)
    lamella = {}
    for n in range(N):
        lamella[n] = ProcessLamella(sample[n], Binaries, Elements, omega, phi,
                                    delpsi, plane, ScanType, Twotheta, lam,
                                    a0, b0, c0, beta0, orient, thetaM)
    # Set up the crystallite tilt (alpha) and azimuth (beta) deviation angles
    # for use in calculating deviation parameter (eta).
    # Additionally, set up the weighting distributions W_alpha and W_beta.
    ang_alpha = np.empty((N, N_alpha))
    ang_beta = np.empty((N, N_beta))
    W_alpha = np.empty((N, N_alpha))
    W_beta = np.empty((N, N_beta))
    for n in range(N):
        sigma_alpha = sample[n]['sigma_alpha']
        sigma_beta = sample[n]['sigma_beta']
        if DistAlpha == 'Gaussian':
            ang_alpha[n, :] = N_sigma*sigma_alpha*np.linspace(-1, 1, N_alpha)
        elif DistAlpha == 'Uniform':
            ang_alpha[n, :] = sigma_alpha*np.linspace(-1, 1, N_alpha)
        else:
            print('Invalid distribution specified for crystallite tilt, must be "Gaussian" or "Uniform"')
        if DistBeta == 'Gaussian':
            ang_beta[n, :] = N_sigma*sigma_beta*np.linspace(-1, 1, N_beta)
        elif DistBeta == 'Uniform':
            ang_beta[n, :] = sigma_beta*np.linspace(-1, 1, N_beta)
        else:
            print('Invalid distribution specified for crystallite azimuth, must be "Gaussian" or "Uniform"')
        W_alpha[n, :] = WeightDist(ang_alpha[n, :], sigma_alpha, N_alpha, DistAlpha)
        W_beta[n, :] = WeightDist(ang_beta[n, :], sigma_beta, N_beta, DistBeta)
    # First calculate deviation parameter eta for the 1st layer (substrate).
    # Calculate solution to Takagi-Taupin equation X using the Darwin-Prins
    # equation (assumes infinite thickness, no dislocations).
    eta = np.zeros((N, N_alpha, N_beta, len(omega)), dtype=complex)
    X = np.zeros((N, N_alpha, N_beta, len(omega)), dtype=complex)
    eta0 = CalcEta(lamella[0], 0, 0)
    for i in range(N_alpha):
        for j in range(N_beta):
            eta[0, i, j, :] = eta0
            X[0, i, j, :] = eta0 - np.sign(np.real(eta0))*np.sqrt(eta0**2 - 1)
    # Calculate unbroadened intensity of substrate.
    I0 = abs(X[0, 0, 0, :])**2
    # Now loop through and calculate eta and diffracted signal X for all
    # lamellae n (except substrate) and all mosaic crystallites i,j.
    # Apply mass attenuation to the diffracted signal amplitude. Assume
    # mass attenuation does not influence signal phase.
    for n in range(1, N):
        theta = lamella[n]['Theta']
        for i in range(N_alpha):
            for j in range(N_beta):
                eta[n, i, j, :] = CalcEta(lamella[n], ang_alpha[n, i], ang_beta[n, j])
                X[n, i, j, :] = CalcX(eta[n, i, j, :], X[n - 1, i, j, :], lamella[n]['T'])
                h = (10**-7)*sample[n]['thickness']
                for m in range(len(theta)):
                    X[n, i, j, m] = X[n, i, j, m]*np.exp(-lamella[n]['mu']*h/(2*np.sin(theta[m])))
    # Calculate the total diffracted intensity from the Nth layer by summing
    # over all the mosaic crystallites with the weighting functions.
    I = np.zeros(len(omega))
    for i in range(N_alpha):
        for j in range(N_beta):
            I += abs(X[-1, i, j, :])**2 * W_alpha[-1, i] * W_beta[-1, j]
    # Specify instrumental broadening full width half maximum (FWHM) in
    # radians. Use the formula from X'Pert Epitaxy help contents, "Default 
    # Convolution Conditions for Simulations". Use the angular divergence and
    # wavelength spread values from the X'Pert Epitaxy sample output files.
    # Assume 2-bounce Ge (220) hybrid monochromator & CuKalpha radiation.
    # Monochromator angle thetaM = 22.6488 degrees for Ge (220).
    thetaC = (max(omega) + min(omega))/2
    dtheta = (np.pi/180) * 0.0026667 * (4/bounces)
    dlam = 0.0004183 * (4/bounces)
    FWHM = (2/2.64)*np.sqrt(dtheta**2 + dlam**2*(np.tan(thetaC) - np.tan(thetaM))**2)
    I = InstrBroaden(I, omega, FWHM)
    I0 = InstrBroaden(I0, omega, FWHM)
    return I, I0, lamella